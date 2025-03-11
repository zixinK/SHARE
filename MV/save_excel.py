import openpyxl
def save_excel(dict_data,d_ir,excel_name):
    keys = list(dict_data.keys())  # 获取字典的key值
    values = list(dict_data.values())  # 获取字典的value值
    # print(keys)
    # print(values)
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('sheet1')
    for key in keys:
        column = keys.index(key)  # 获取key值的索引
        row_value = values[column]  # 根据获取到的key值索引来获取对应的value值，此时的value值还是一个list类型的
        ws.cell(row=1, column=column + 1).value = key  # 把获取的每一个key值作为列，放在excel表中的第一行
        for i in range(len(row_value)):
            row = i + 2 # 获取索引，+2是因为要写入excel表格，从第二行开始写入，因为第一行作为列已经使用了。
            ws.cell(row=row, column=column + 1).value = row_value[i]  # 在每个列的行下面写入数据，例如在第一列，第二行、第三行、第四行一次写入数据。
    wb.save(f'{d_ir}/{excel_name}.xlsx')   #你要保存的excel数据
